import openpyxl
import requests
import json
from secret_folder.secret_data import token_novaposhta as token

file = \
'/Users/stanislav_zhucenko/PycharmProjects/excel_automation/test_data/Реєстр платежів контрагента Ошега Н.А. ФОП № 5226024 від 13 грудня 2024 р.XLSX'
workbook_obj = openpyxl.load_workbook(filename=file)

sheet_obj = workbook_obj.active

input_date = '13.12.2024'  # Choose the date of change in the state of money transfer
ttn_obj = sheet_obj['I']  # The column with TTN
# amount_obj = sheet_obj['D']  # The column with the amount of money to the TTN

# Перебор всех листов в книге
for sheet in workbook_obj.worksheets:
    # Создаем копию списка объединённых ячеек, так как нельзя изменять список во время итерации
    merged_cells = list(sheet.merged_cells)

    # Убираем объединение для каждой области
    for merged_range in merged_cells:
        sheet.unmerge_cells(str(merged_range))

for sheet in workbook_obj.worksheets:
    # Сброс ширины всех столбцов
    for col in sheet.column_dimensions.values():
        col.hidden = None  # Remove hidden property
        col.outline_level = 0  # Reset outline level
        col.width = 20  # Вернуть стандартное значение (можно указать число, например, 10)

    # Сброс высоты всех строк
    for row in sheet.row_dimensions.values():
        row.hidden = None  # Remove hidden property
        row.outline_level = 0  # Reset outline level
        row.height = None  # Вернуть стандартное значение (можно указать число, например, 15)

sheet_obj.delete_rows(1, amount=12)

def novaposhta_get_barcode(list_of_ttn):
    """
    :param list_of_ttn: list of data with dict with keys "DocumentNumber" and "Phone"(optional)
    :return: dict with keys "DocumentNumber" and values "ClientBarcode"
    """
    url = 'https://api.novaposhta.ua/v2.0/json/'

    data = {
        "apiKey": token,
        "modelName": "TrackingDocumentGeneral",
        "calledMethod": "getStatusDocuments",
        "methodProperties": {
            "Documents": list_of_ttn

        }
    }

    json_data = json.dumps(data, indent=4, ensure_ascii=False)

    response = requests.post(url=url, data=json_data)

    parse_data = json.loads(response.content.decode('utf-8'))

    ttn_barcode = dict()

    for el in parse_data['data']:
        ttn_barcode[el['Number']] = el['AdditionalInformationEW']

    return ttn_barcode


filename = f"{input_date}_Novapay.xlsx"
#
#
workbook_obj.save(filename=filename)

workbook_obj = openpyxl.load_workbook(filename=filename)

sheet_obj = workbook_obj.active
ttn_obj = sheet_obj['I']  # The column with TTN
add_data = []
for ttn in ttn_obj:
    # add_data.append(novaposhta_get_barcode([str(ttn._value), ]))
    if ttn.value is not None:
        value_to_set = novaposhta_get_barcode([str(ttn.value), ])[str(ttn.value)]
        sheet_obj.cell(row=ttn.row, column=15, value=value_to_set)
# print(add_data)
filename = f"{input_date}_Novapay.xlsx"
#
#
workbook_obj.save(filename=filename)


# print(f"Your file {filename} is already done!")
